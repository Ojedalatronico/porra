import os
import csv
import json
import requests
from openpyxl import load_workbook

CACHE_FILE = os.path.join("data", "predictions_cache.json")

def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

def get_real_results():
    """Fetches real match results from the API."""
    print("Fetching real results from API...")
    try:
        response = requests.get("https://worldcup26.ir/get/games", timeout=10)
        if response.status_code == 200:
            return response.json()
        print(f"API Error: Status {response.status_code}")
    except Exception as e:
        print(f"Error consulting API: {e}")
    return None

def save_api_results(api_results, output_csv):
    """Saves API results to a separate CSV file, formatted identically to participants' data."""
    if not api_results or "games" not in api_results:
        return

    print(f"Saving API results to {output_csv}...")
    try:
        parsed_results = []
        for i, game in enumerate(api_results["games"], 1):
            # Use API id or index as fallback
            game_id = game.get("id", str(i))
            parsed_results.append({
                "ID": game_id,
                "Participante": "REAL_RESULTS",
                "Tipo": game.get("type", "group").capitalize(),
                "Equipo_Local": game.get("home_team_name_en"),
                "Goles_Local": game.get("home_score"),
                "Goles_Visitante": game.get("away_score"),
                "Equipo_Visitante": game.get("away_team_name_en")
            })

        keys = ["ID", "Participante", "Tipo", "Equipo_Local", "Goles_Local", "Goles_Visitante", "Equipo_Visitante"]
        output_path = os.path.join("data", output_csv)
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            dict_writer = csv.DictWriter(f, fieldnames=keys)
            dict_writer.writeheader()
            dict_writer.writerows(parsed_results)
            
        print("API results saved successfully!")
    except Exception as e:
        print(f"Error saving API results: {e}")

def calculate_points(pred, real_match, real_teams_by_phase):
    """Calculates points for a row based on predefined rules."""
    if not real_match:
        return 0
    
    # If the match hasn't finished, no points are counted yet
    if real_match.get("finished") != "TRUE":
        return 0
    
    tipo = pred["Tipo"]
    try:
        p_l = int(pred["Goles_Local"]) if pred["Goles_Local"] is not None else 0
        p_v = int(pred["Goles_Visitante"]) if pred["Goles_Visitante"] is not None else 0
        r_l = int(real_match["Goles_Local"]) if real_match["Goles_Local"] is not None else 0
        r_v = int(real_match["Goles_Visitante"]) if real_match["Goles_Visitante"] is not None else 0
    except (ValueError, TypeError):
        return 0

    # 1. Group Stage
    if tipo == "Fase de Grupos":
        if p_l == r_l and p_v == r_v:
            return 3
        # Sign: 1 (home win), -1 (away win), 0 (draw)
        sig_p = (p_l > p_v) - (p_l < p_v)
        sig_r = (r_l > r_v) - (r_l < r_v)
        if sig_p == sig_r:
            return 1
        return 0

    # 2. Knockout Rounds (Points for teams qualifying)
    points = 0
    phase_map = {
        "Dieciseisavos": 3, "Round of 32": 3, "1/16": 3,
        "Octavos": 4, "Round of 16": 4, "1/8": 4,
        "Cuartos": 5, "Quarter-finals": 5, "1/4": 5,
        "Semifinales": 6, "Semi-finals": 6, "1/2": 6,
        "Final": 7
    }
    
    points_per_team = 0
    for key, val in phase_map.items():
        if key.lower() in tipo.lower():
            points_per_team = val
            break
    
    if points_per_team > 0:
        # Check if teams actually reached this phase in reality
        actual_teams_in_phase = real_teams_by_phase.get(tipo, set())
        if not actual_teams_in_phase:
            # Retry with partial match if phase name mapping varies
            for phase_name, teams in real_teams_by_phase.items():
                if any(k.lower() in phase_name.lower() for k in phase_map.keys() if k.lower() in tipo.lower()):
                    actual_teams_in_phase = teams
                    break

        if pred["Equipo_Local"] in actual_teams_in_phase:
            points += points_per_team
        if pred["Equipo_Visitante"] in actual_teams_in_phase:
            points += points_per_team

    # 3. Final Winner (8 extra points)
    if "final" in tipo.lower() and "quarter" not in tipo.lower():
        # Predicted winner
        pred_winner = pred["Equipo_Local"] if p_l > p_v else pred["Equipo_Visitante"]
        # Actual winner (determined by goals)
        real_winner = real_match["Equipo_Local"] if r_l > r_v else real_match["Equipo_Visitante"]
        if pred_winner == real_winner and pred_winner is not None:
            points += 8

    return points

def parse_excel_files(data_folder, output_csv, api_csv):
    """Parses Excel files using a cache to avoid re-processing unchanged files."""
    files = sorted([f for f in os.listdir(data_folder) if f.endswith(".xlsx")])
    
    # Load existing cache
    cache = load_cache()
    new_cache = {}
    all_data = []

    # Get real results for points calculation
    api_raw = get_real_results()
    real_matches_lookup = {}
    real_teams_by_phase = {}

    if api_raw and "games" in api_raw:
        save_api_results(api_raw, api_csv)
        for game in api_raw["games"]:
            g_id = str(game.get("id"))
            real_matches_lookup[g_id] = {
                "Goles_Local": game.get("home_score"),
                "Goles_Visitante": game.get("away_score"),
                "Equipo_Local": game.get("home_team_name_en"),
                "Equipo_Visitante": game.get("away_team_name_en"),
                "finished": game.get("finished")
            }
            tipo_api = game.get("type", "group")
            if tipo_api not in real_teams_by_phase:
                real_teams_by_phase[tipo_api] = set()
            real_teams_by_phase[tipo_api].add(game.get("home_team_name_en"))
            real_teams_by_phase[tipo_api].add(game.get("away_team_name_en"))

    # Process each file
    for filename in files:
        file_path = os.path.join(data_folder, filename)
        mtime = os.path.getmtime(file_path)
        
        # Determine participant name
        if "2026_" in filename:
            participant_name = filename.split("2026_")[1].replace(".xlsx", "").strip()
        else:
            participant_name = filename.replace(".xlsx", "").strip()

        # Check if we can use cache
        if filename in cache and cache[filename]["mtime"] == mtime:
            print(f"Using cache for: {participant_name}")
            raw_predictions = cache[filename]["raw_predictions"]
        else:
            print(f"Parsing Excel: {participant_name}...")
            raw_predictions = []
            match_counter = 1
            try:
                wb = load_workbook(file_path, data_only=True)
                if "2026 World Cup" in wb.sheetnames:
                    ws = wb["2026 World Cup"]
                    
                    # 1. Group Stage
                    for row in ws.iter_rows(min_row=7, max_row=78, min_col=5, max_col=8):
                        if row[0].value:
                            raw_predictions.append({
                                "ID": str(match_counter),
                                "Tipo": "Fase de Grupos",
                                "Equipo_Local": str(row[0].value).strip(),
                                "Goles_Local": row[1].value,
                                "Goles_Visitante": row[2].value,
                                "Equipo_Visitante": str(row[3].value).strip()
                            })
                            match_counter += 1

                    # 2. Knockout Stages
                    current_phase = "Eliminatorias"
                    for col in range(63, 96):
                        phase_header = ws.cell(row=6, column=col).value
                        if phase_header: current_phase = str(phase_header).strip()
                        r = 8
                        while r < 71:
                            v1 = ws.cell(row=r, column=col).value
                            v2 = ws.cell(row=r+1, column=col).value
                            if isinstance(v1, str) and isinstance(v2, str) and len(v1) > 2:
                                raw_predictions.append({
                                    "ID": str(match_counter),
                                    "Tipo": current_phase,
                                    "Equipo_Local": v1.strip(),
                                    "Goles_Local": ws.cell(row=r, column=col+1).value,
                                    "Goles_Visitante": ws.cell(row=r+1, column=col+1).value,
                                    "Equipo_Visitante": v2.strip()
                                })
                                match_counter += 1
                                r += 2
                            else: r += 1

                    # 3. Final
                    eq1 = ws.cell(row=37, column=92).value
                    eq2 = ws.cell(row=38, column=92).value
                    if eq1 and eq2:
                        raw_predictions.append({
                            "ID": str(match_counter),
                            "Tipo": "Final",
                            "Equipo_Local": str(eq1).strip(),
                            "Goles_Local": ws.cell(row=37, column=93).value,
                            "Goles_Visitante": ws.cell(row=38, column=93).value,
                            "Equipo_Visitante": str(eq2).strip()
                        })
                wb.close()
            except Exception as e:
                print(f"Error parsing {filename}: {e}")
        
        # Update current cache and calculate points for this run
        new_cache[filename] = {
            "mtime": mtime,
            "raw_predictions": raw_predictions
        }
        
        for pred in raw_predictions:
            final_row = pred.copy()
            final_row["Participante"] = participant_name
            final_row["Puntos"] = calculate_points(pred, real_matches_lookup.get(pred["ID"]), real_teams_by_phase)
            all_data.append(final_row)

    # Save cache for next time
    save_cache(new_cache)

    # Save to CSV
    keys = ["ID", "Participante", "Tipo", "Equipo_Local", "Goles_Local", "Goles_Visitante", "Equipo_Visitante", "Puntos"]
    output_path = os.path.join("data", output_csv)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        dict_writer = csv.DictWriter(f, fieldnames=keys)
        dict_writer.writeheader()
        dict_writer.writerows(all_data)
    
    print(f"\nDone! Updated {len(all_data)} rows. Saved to {output_path}")

if __name__ == "__main__":
    # Ensure script runs correctly from project root
    parse_excel_files("data", "resultados_porra.csv", "resultados_reales.csv")
